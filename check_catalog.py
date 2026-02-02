import os
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import PatternFill

import pymorphy3
import language_tool_python
import httpx
from dotenv import load_dotenv

# Загружаем переменные окружения из .env
load_dotenv()

# Импортируем настройку логирования
from logging_config import setup_logging

# Настраиваем логирование
logger = setup_logging()


"""
Инструмент для проверки выгрузки каталога:
- вход: CSV `AZ_Игрушки_ru_RU_2025-11-18.csv` (разделитель `;`);
- выход: Excel `AZ_Игрушки_ru_RU_2025-11-18_checked.xlsx`:
  - те же данные;
  - для проверяемых текстовых колонок добавлены:
    - `<col>__correct` — предложенное исправление;
    - `<col>__comment` — комментарии по типу ошибки;
  - исходные ячейки с ошибками подсвечены цветом.

РЕАЛИЗОВАНЫ ПРАВИЛА:
- Грамматика/орфография (LanguageTool) для:
  - русских названий категорий (все уровни);
  - названий параметров;
  - значений параметров.
- Категории:
  - должны быть во множественном числе (кроме списка исключений «неисчисляемых»/неконкретных слов).
- Значения параметров:
  - паттерн «Другой/Другое/Другая/Другие + название параметра» с согласованием прилагательного «другой»
    по роду/числу названия параметра (через pymorphy3);
  - единый формат регистра (первая буква строчная для обычных прилагательных/существительных);
  - единообразие части речи внутри одного параметра (по `param_id`):
    все значения стремимся привести к преобладающей части речи (прилагательное/существительное),
    пока только подсветкой и комментариями.

УТОЧНЕНИЯ:
- Для «штучности товара» отдельной явной колонки нет, поэтому:
  - мы предполагаем, что значения, описывающие характеристики товара (например, цвет),
    должны быть в форме ЕДИНСТВЕННОГО числа (эвристики по части речи/окончаниям).
"""


CONFIG = {
    "input_file": "AZ_Игрушки_ru_RU_2025-11-18.csv",
    "output_file": "AZ_Игрушки_ru_RU_2025-11-18_checked.xlsx",
    "sep": ";",
    # Файл выгрузки, судя по ошибке, не в UTF-8. Пробуем стандартную Windows‑кодировку для русского.
    "encoding": "cp1251",
    "category_columns": [
        "category_level_1_name",
        "category_level_2_name",
        "category_level_3_name",
        "category_level_4_name",
        "category_level_5_name",
    ],
    "param_name_column": "param_name",
    "param_value_column": "value_name",
    "param_id_column": "param_id",
    # Порог для определения преобладающего паттерна регистра (60% по умолчанию)
    "case_consistency_threshold": 0.6,
}


_MORPH = pymorphy3.MorphAnalyzer(lang="ru")
try:
    _LT = language_tool_python.LanguageTool("ru-RU")
except Exception:
    # Если нет Java или LanguageTool не может стартовать — отключаем орфографию,
    # но не ломаем весь скрипт.
    _LT = None
    print(
        "Внимание: не удалось запустить LanguageTool (скорее всего, не установлен Java). "
        "Проверки орфографии/грамматики будут пропущены."
    )


# --- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ: МОРФОЛОГИЯ И ГРАММАТИКА ---


def _first_parse(word: str):
    parses = _MORPH.parse(word)
    return parses[0] if parses else None


def check_spelling(text: str) -> Optional[Tuple[str, str]]:
    """
    Проверка орфографии/грамматики через LanguageTool.
    Возвращает (как_было, как_нужно) с первой авто-подсказкой, если есть.
    """
    if _LT is None:
        return None

    text = (text or "").strip()
    if not text:
        return None

    matches = _LT.check(text)
    if not matches:
        return None

    first = matches[0]
    if not first.replacements:
        return None

    start = first.offset
    end = start + first.errorLength
    replacement = first.replacements[0]
    corrected = text[:start] + replacement + text[end:]
    if corrected != text:
        return text, corrected
    return None


def is_plural_noun(word: str) -> bool:
    """
    Проверка, что слово — существительное во множественном числе.
    Проверяет все возможные разборы слова, т.к. "машинки" может быть:
    - ед.ч., род.падеж (машинка -> машинки)
    - мн.ч., им.падеж (машинки)
    """
    word = (word or "").strip()
    if not word:
        return False

    parses = _MORPH.parse(word)
    if not parses:
        return False

    # Проверяем все возможные разборы
    for p in parses:
        # Ищем форму: существительное + множественное число + именительный падеж
        if "NOUN" in p.tag and "plur" in p.tag and "nomn" in p.tag:
            return True

    return False


def make_plural(word: str) -> Optional[str]:
    """
    Предложить форму во множественном числе, если возможно.
    """
    word = (word or "").strip()
    if not word:
        return None

    p = _first_parse(word)
    if not p:
        return None

    form = p.inflect({"plur"})
    if not form:
        return None

    w = form.word
    if word[0].isupper():
        w = w.capitalize()
    return w


_CATEGORY_MASS_LIKE = {
    # слова, которые в контексте каталога считаем "массовыми"/без реального множественного числа
    "клей",
    "молоко",
    "масло",
    "сахар",
    "транспорт",  # по твоему правилу: "транспорт" не имеет множественного числа
    "игрушечный транспорт",
}


# Список предлогов и союзов для определения составных названий
_FUNCTION_WORDS = {"и", "в", "на", "с", "к", "по", "о", "для", "от", "до", "из", "у", "при"}

# Известные бренды, названия мест и устойчивые выражения (имена собственные)
_KNOWN_PROPER_NOUNS = {
    "детский мир",
    "красная площадь",
    "чёрное море",
    "белое море",
    "золотое кольцо",
    "третьяковская галерея",
}


def is_proper_noun_or_compound(name: str) -> bool:
    """
    Определяет, является ли название категории именем собственным или устойчивым составным выражением,
    которое не следует преобразовывать во множественное число.

    Эвристики:
    1. Проверка в списке известных имён собственных
    2. Содержит латиницу или цифры (бренд)
    3. Все значимые слова начинаются с заглавной буквы (имя собственное)
    4. Паттерн "Прилагательное + Существительное" с заглавными буквами
    """
    name = (name or "").strip()
    if not name:
        return False

    # 1. Проверка в списке известных имён собственных
    if name.lower() in _KNOWN_PROPER_NOUNS:
        return True

    # 2. Если есть латиница или цифры — вероятно бренд
    if any(ch.isascii() and ch.isalpha() for ch in name) or any(ch.isdigit() for ch in name):
        return True

    # 3. Многословное название — проверяем паттерн заглавных букв
    words = name.split()
    if len(words) >= 2:
        # Проверяем все значимые слова (не предлоги/союзы)
        significant_words = [w for w in words if w.lower() not in _FUNCTION_WORDS]

        if len(significant_words) >= 2:
            # Проверка паттерна "Прилагательное(ые) + Существительное"
            # Составные названия категорий типа "Игрушечные машинки", "Другой игрушечный транспорт"
            # не нужно склонять
            last_word = significant_words[-1]
            last_parse = _first_parse(last_word)

            # Проверяем что последнее слово - существительное
            if last_parse and "NOUN" in last_parse.tag:
                # Проверяем что хотя бы одно из предыдущих слов - прилагательное
                has_adjective = False
                for word in significant_words[:-1]:
                    word_parse = _first_parse(word)
                    if word_parse and ("ADJF" in word_parse.tag or "ADJS" in word_parse.tag):
                        has_adjective = True
                        break

                if has_adjective:
                    # Составное название категории (прилагательное(ые) + существительное)
                    return True

            # Если все значимые слова с заглавной — вероятно имя собственное
            all_capitalized = all(w[0].isupper() for w in significant_words if w)
            if all_capitalized:
                return True

    # 3. Проверка через pymorphy3 на тег Name (имя собственное)
    # Проверяем первое слово
    first_word = words[0] if words else name
    parsed = _first_parse(first_word)
    if parsed and "Name" in parsed.tag:
        return True

    return False


def ensure_category_format(name: str) -> Optional[Tuple[str, str]]:
    """
    Проверка формата названия категории по спецификации:
    1. Первая буква заглавная (кроме аббревиатур/брендов)
    2. Именительный падеж (nominative)
    3. Множественное число (кроме неисчисляемых и составных названий)

    Примеры:
    - "игрушка" -> "Игрушки"
    - "игрушек" -> "Игрушки"
    - "USB кабели" -> "USB кабели" (не меняем)
    - "Игрушечный транспорт" -> "Игрушечный транспорт" (составное название, не меняем)
    """
    name = (name or "").strip()
    if not name:
        return None

    # Проверка списка исключений (неисчисляемые)
    low = name.lower()
    if low in _CATEGORY_MASS_LIKE:
        return None

    # Проверка имён собственных, брендов и составных названий (не трогаем)
    if is_proper_noun_or_compound(name):
        return None

    # Проверяем, является ли это аббревиатурой (все буквы заглавные, > 1 буквы)
    words = name.split()
    is_abbreviation = len(name) > 1 and name.replace(" ", "").isupper() and name.replace(" ", "").isalpha()

    if is_abbreviation:
        return None  # Не трогаем аббревиатуры типа "USB", "DVD"

    # Анализируем последнее слово (главное существительное в составных названиях)
    main_word = words[-1] if len(words) > 1 else name
    prefix = " ".join(words[:-1]) + " " if len(words) > 1 else ""

    # Парсим главное слово
    parsed = _first_parse(main_word)
    if not parsed or "NOUN" not in parsed.tag:
        return None  # Не существительное - не трогаем

    # Проверяем текущее состояние
    is_plural = "plur" in parsed.tag
    is_nominative = "nomn" in parsed.tag
    is_capitalized = main_word[0].isupper() if main_word else False

    # Если уже в правильной форме - не трогаем
    if is_plural and is_nominative and is_capitalized:
        return None

    # Формируем правильную форму: именительный падеж + мн.число
    try:
        # Приводим к им.падежу, мн.числу
        correct_form = parsed.inflect({"nomn", "plur"})
        if not correct_form:
            return None

        corrected = correct_form.word

        # Первая буква заглавная
        if corrected and corrected[0].islower():
            corrected = corrected[0].upper() + corrected[1:]

        # Собираем полное название с префиксом
        full_corrected = prefix + corrected

        if full_corrected != name:
            return name, full_corrected

    except Exception:
        return None

    return None


# Оставляем старую функцию для обратной совместимости, но она теперь вызывает новую
def ensure_category_plural(name: str) -> Optional[Tuple[str, str]]:
    """Устаревшая функция - используйте ensure_category_format"""
    return ensure_category_format(name)


def detect_pos(word: str) -> Optional[str]:
    """
    Определить часть речи: возвращает "ADJ", "NOUN" или None.
    """
    word = (word or "").strip()
    if not word:
        return None

    p = _first_parse(word)
    if not p:
        return None

    if "ADJF" in p.tag or "ADJS" in p.tag:
        return "ADJ"
    if "NOUN" in p.tag:
        return "NOUN"
    return None


def is_proper_noun_word(word: str) -> bool:
    """
    Check if word should stay capitalized (abbreviation/brand/proper noun).
    - All uppercase with length > 1: USB, DVD, BMW
    - Contains Latin letters: iPhone, MacBook
    - Mixed case with Latin: iOS, PlayStation
    """
    if not word:
        return False

    # All uppercase abbreviations (USB, DVD, BMW)
    if word.isupper() and len(word) > 1:
        return True

    # Contains Latin letters (brands: iPhone, PlayStation, etc.)
    if any(ch.isascii() and ch.isalpha() for ch in word):
        return True

    return False


def normalize_compound_capitalization(text: str) -> str:
    """
    Capitalize first word, lowercase rest unless proper noun/abbreviation.
    Also handles hyphens and conjunction "и".

    Examples:
        "Другой Красный Цвет" → "Другой красный цвет"
        "Игрушки-Роботы" → "Игрушки-роботы"
        "Роботы и Трансформеры" → "Роботы и трансформеры"
        "Другой iPhone" → "Другой iPhone"
        "Другая USB мышь" → "Другая USB мышь"
    """
    if not text or not text.strip():
        return text

    # Split by spaces first
    words = text.split()
    if len(words) <= 1:
        # Check if single word contains hyphen
        if '-' in text:
            parts = text.split('-')
            result_parts = [parts[0].capitalize()]
            for part in parts[1:]:
                if is_proper_noun_word(part):
                    result_parts.append(part)
                else:
                    result_parts.append(part.lower())
            return '-'.join(result_parts)
        return text

    result = []

    for i, word in enumerate(words):
        # Handle hyphenated words
        if '-' in word:
            parts = word.split('-')
            if i == 0:
                # First word: capitalize first part, lowercase rest
                processed_parts = [parts[0].capitalize()]
                for part in parts[1:]:
                    if is_proper_noun_word(part):
                        processed_parts.append(part)
                    else:
                        processed_parts.append(part.lower())
            else:
                # Not first word: lowercase all parts unless proper noun
                processed_parts = []
                for part in parts:
                    if is_proper_noun_word(part):
                        processed_parts.append(part)
                    else:
                        processed_parts.append(part.lower())
            result.append('-'.join(processed_parts))
        else:
            # Regular word without hyphen
            if i == 0:
                result.append(word.capitalize())  # First word always capitalized
            elif word.lower() == 'и':
                result.append('и')  # Conjunction stays lowercase
            elif is_proper_noun_word(word):
                result.append(word)  # Keep proper nouns as-is
            else:
                result.append(word.lower())  # Lowercase everything else

    return " ".join(result)


def singularize_noun(word: str) -> str:
    """
    Convert noun to singular form (nominative case).

    Examples:
        "особенности" → "особенность"
        "цвета" → "цвет"
        "формы" → "форма"
    """
    if not word or not word.strip():
        return word

    p = _first_parse(word.strip())
    if not p:
        return word

    # If already singular, return as-is
    if "sing" in p.tag:
        return word

    # If plural, try to inflect to singular nominative
    if "plur" in p.tag:
        try:
            singular_form = p.inflect({"sing", "nomn"})
            if singular_form:
                return singular_form.word
        except Exception:
            pass

    return word


def extract_head_noun(phrase: str) -> str:
    """
    Extract grammatical head (main noun) from a phrase.
    In Russian, the head noun is typically the LAST noun in the phrase.

    Examples:
        "Цвет корпуса" → "корпуса" (genitive of "корпус")
        "Материал изготовления" → "изготовления"
        "Бренд" → "Бренд"
    """
    words = phrase.split()
    if not words:
        return phrase

    # In Russian, head noun is typically LAST noun in phrase
    for word in reversed(words):
        p = _first_parse(word)
        if p and "NOUN" in p.tag:
            return word

    # Fallback to first word
    return words[0]


def normalize_other_pattern(param_name: str, value: str) -> Optional[Tuple[str, str]]:
    """
    Паттерн «другой/другое/другая/другие ... + <название параметра>»:
    - базовый шаблон: прилагательное "другой" + param_name;
    - форму "другой" подбираем по падежу/роду/числу param_name (грубо по слову в именительном).
    - param_name приводится к единственному числу.
    """
    value = (value or "").strip()
    param_name = (param_name or "").strip()
    if not value or not param_name:
        return None

    low = value.lower()
    if not low.startswith("друг"):
        return None

    # Приводим параметр к единственному числу
    # "особенности" → "особенность"
    param_name_singular = " ".join(singularize_noun(word) for word in param_name.split())

    # Морфологический разбор названия параметра
    # Берём главное (головное) существительное из фразы
    head = extract_head_noun(param_name_singular)
    p = _first_parse(head)
    if not p:
        correct = f"Другой {param_name}"
        return (value, correct) if value != correct else None

    tags = p.tag
    grammemes = set()
    # род
    if "masc" in tags:
        grammemes.add("masc")
    if "femn" in tags:
        grammemes.add("femn")
    if "neut" in tags:
        grammemes.add("neut")
    # число: по умолчанию единственное
    if "plur" in tags:
        grammemes.add("plur")
    else:
        grammemes.add("sing")
    # падеж — именительный
    grammemes.add("nomn")

    # "другой" в нужной форме
    base = _first_parse("другой")
    if base:
        inflected = base.inflect(grammemes)
        if inflected:
            other_word = inflected.word.capitalize()
        else:
            other_word = "Другой"
    else:
        other_word = "Другой"

    # Применяем правильную капитализацию к составной фразе (используем единственное число)
    correct = normalize_compound_capitalization(f"{other_word} {param_name_singular}")
    if value != correct:
        return value, correct
    return None


def check_case_consistency(value: str, expected_case: str) -> Optional[Tuple[str, str]]:
    """
    Проверка единообразия регистра значения параметра относительно ожидаемого паттерна.
    expected_case: "lowercase" или "uppercase"
    """
    value = (value or "").strip()
    if not value or not value[0].isalpha():
        return None

    # Пропускаем слова с латиницей/цифрами (вероятно бренды)
    if any(ch.isascii() and ch.isalpha() for ch in value) or any(ch.isdigit() for ch in value):
        return None

    actual_case = "uppercase" if value[0].isupper() else "lowercase"

    if actual_case != expected_case:
        if expected_case == "lowercase":
            corrected = value[0].lower() + value[1:]
        else:
            corrected = value[0].upper() + value[1:]

        if corrected != value:
            return value, corrected

    return None


def normalize_value_format(value: str) -> Optional[Tuple[str, str]]:
    """
    УСТАРЕВШАЯ ФУНКЦИЯ: Заменена на check_case_consistency.
    Оставлена для обратной совместимости, но не используется в основном коде.

    Формат значений:
    - первая буква строчная (если это обычное слово, а не бренд/аббревиатура).
    """
    value = (value or "").strip()
    if not value:
        return None

    # если в слове есть латиница или цифры — скорее всего бренд, не трогаем
    if any(ch.isascii() and ch.isalpha() for ch in value) or any(ch.isdigit() for ch in value):
        return None

    if value[0].isupper() and value[0].isalpha():
        corrected = value[0].lower() + value[1:]
        if corrected != value:
            return value, corrected
    return None


def ensure_singular_for_item(value: str) -> Optional[Tuple[str, str]]:
    """
    Эвристика: значения, описывающие одну вещь (цвет, материал и т.п.),
    должны быть в форме единственного числа.
    """
    value = (value or "").strip()
    if not value:
        return None

    p = _first_parse(value)
    if not p:
        return None

    # Проверяем, что слово действительно в именительном падеже множественного числа
    # (не путаем с другими падежами единственного числа)
    if "plur" not in p.tag or "nomn" not in p.tag:
        return None

    # Пробуем получить форму единственного числа
    form = p.inflect({"sing", "nomn"}) if "NOUN" in p.tag or "ADJF" in p.tag or "ADJS" in p.tag else None
    if not form:
        return None

    corrected = form.word
    if value[0].isupper():
        corrected = corrected.capitalize()

    # Проверяем, что форма действительно изменилась и это не ложное срабатывание
    if corrected != value and corrected.lower() != value.lower():
        # Дополнительная проверка: слово должно действительно иметь множественное число
        # Пробуем проверить обратное преобразование
        check_plural = _MORPH.parse(corrected)[0].inflect({"plur", "nomn"}) if _MORPH.parse(corrected) else None
        if check_plural and check_plural.word.lower() == value.lower():
            # Подтверждено: это действительно множественное число
            return value, corrected

    return None


# --- ОСНОВНАЯ ЛОГИКА ---


def build_param_pos_profile(df: pd.DataFrame) -> Dict[str, str]:
    """
    Для каждого param_id определяет преобладающую часть речи его значений (ADJ / NOUN).
    Это нужно, чтобы проверять однородность формата значений параметра.
    """
    cfg = CONFIG
    pid_col = cfg["param_id_column"]
    val_col = cfg["param_value_column"]

    if pid_col not in df.columns or val_col not in df.columns:
        return {}

    stats: Dict[str, Dict[str, int]] = {}

    for _, row in df.iterrows():
        pid = row.get(pid_col)
        val = row.get(val_col)
        if pid is None or pd.isna(pid) or val is None or pd.isna(val):
            continue
        pid = str(pid)
        word = str(val).strip()
        if not word:
            continue

        pos = detect_pos(word)
        if not pos:
            continue

        if pid not in stats:
            stats[pid] = {"ADJ": 0, "NOUN": 0}
        stats[pid][pos] += 1

    result: Dict[str, str] = {}
    for pid, counts in stats.items():
        if counts["ADJ"] == 0 and counts["NOUN"] == 0:
            continue
        result[pid] = "ADJ" if counts["ADJ"] >= counts["NOUN"] else "NOUN"
    return result


def build_param_case_profile(df: pd.DataFrame) -> Dict[str, str]:
    """
    Для каждого param_id определяет преобладающий паттерн регистра его значений (lowercase / uppercase).
    Это нужно, чтобы проверять единообразие регистра внутри группы значений параметра.
    """
    cfg = CONFIG
    pid_col = cfg["param_id_column"]
    val_col = cfg["param_value_column"]
    threshold = cfg.get("case_consistency_threshold", 0.6)

    if pid_col not in df.columns or val_col not in df.columns:
        return {}

    stats: Dict[str, Dict[str, int]] = {}

    for _, row in df.iterrows():
        pid = row.get(pid_col)
        val = row.get(val_col)
        if pid is None or pd.isna(pid) or val is None or pd.isna(val):
            continue
        pid = str(pid)
        word = str(val).strip()
        if not word or not word[0].isalpha():
            continue

        # Пропускаем слова с латиницей/цифрами (вероятно бренды)
        if any(ch.isascii() and ch.isalpha() for ch in word) or any(ch.isdigit() for ch in word):
            continue

        case_type = "uppercase" if word[0].isupper() else "lowercase"

        if pid not in stats:
            stats[pid] = {"lowercase": 0, "uppercase": 0}
        stats[pid][case_type] += 1

    result: Dict[str, str] = {}
    for pid, counts in stats.items():
        total = counts["lowercase"] + counts["uppercase"]
        if total == 0:
            continue

        lowercase_ratio = counts["lowercase"] / total
        uppercase_ratio = counts["uppercase"] / total

        # Определяем преобладающий паттерн только если есть явное большинство
        if lowercase_ratio >= threshold:
            result[pid] = "lowercase"
        elif uppercase_ratio >= threshold:
            result[pid] = "uppercase"
        # Если нет явного большинства — не добавляем в профиль

    return result


def has_preposition(text: str) -> bool:
    """
    Проверяет, содержит ли текст предлоги (в, на, из, к, с, у, о, по, до, для и т.д.).
    Такие фразы не должны проверяться на число.
    """
    prepositions = {'в', 'на', 'из', 'к', 'с', 'у', 'о', 'об', 'от', 'до', 'по', 'для', 'без', 'под', 'над', 'при', 'про', 'через', 'за', 'перед', 'между'}
    words = text.lower().split()
    return len(words) > 1 and any(word in prepositions for word in words)


def build_param_number_profile(df: pd.DataFrame) -> Dict[str, str]:
    """
    Для каждого param_id определяет преобладающий паттерн числа его значений (singular / plural).
    Это нужно, чтобы проверять единообразие числа внутри группы значений параметра.
    """
    cfg = CONFIG
    pid_col = cfg["param_id_column"]
    val_col = cfg["param_value_column"]
    threshold = cfg.get("case_consistency_threshold", 0.6)

    if pid_col not in df.columns or val_col not in df.columns:
        return {}

    stats: Dict[str, Dict[str, int]] = {}

    for _, row in df.iterrows():
        pid = row.get(pid_col)
        val = row.get(val_col)
        if pid is None or pd.isna(pid) or val is None or pd.isna(val):
            continue
        pid = str(pid)
        word = str(val).strip()
        if not word:
            continue

        # Пропускаем фразы с предлогами (например "Доставка в район")
        if has_preposition(word):
            continue

        # Пропускаем слова с латиницей/цифрами (вероятно бренды)
        if any(ch.isascii() and ch.isalpha() for ch in word) or any(ch.isdigit() for ch in word):
            continue

        parsed = _first_parse(word)
        if not parsed:
            continue

        # Определяем число
        if "plur" in parsed.tag and "nomn" in parsed.tag:
            number_type = "plural"
        elif "sing" in parsed.tag or "nomn" in parsed.tag:
            number_type = "singular"
        else:
            continue

        if pid not in stats:
            stats[pid] = {"singular": 0, "plural": 0}
        stats[pid][number_type] += 1

    result: Dict[str, str] = {}
    for pid, counts in stats.items():
        total = counts["singular"] + counts["plural"]
        if total == 0:
            continue

        singular_ratio = counts["singular"] / total
        plural_ratio = counts["plural"] / total

        # Определяем преобладающий паттерн только если есть явное большинство
        if singular_ratio >= threshold:
            result[pid] = "singular"
        elif plural_ratio >= threshold:
            result[pid] = "plural"
        # Если нет явного большинства — не добавляем в профиль

    return result


SEMANTIC_URL = os.environ.get("SEMANTIC_URL", "")

# Кэш для LLM результатов (чтобы не проверять одинаковые категории дважды)
_LLM_CACHE: Dict[str, Optional[Tuple[str, str]]] = {}


def semantic_category_suggestion(name: str, path: str) -> Optional[Tuple[str, str]]:
    """
    Запрашивает у LLM-сервиса рекомендацию по названию категории.
    Возвращает (как_было, как_нужно) или None.
    С кэшированием - не проверяет одинаковые категории дважды.
    """
    if not SEMANTIC_URL:
        return None

    name = (name or "").strip()
    path = (path or "").strip()
    if not name:
        return None

    # Проверяем кэш
    cache_key = name.lower()
    if cache_key in _LLM_CACHE:
        return _LLM_CACHE[cache_key]

    try:
        resp = httpx.post(
            SEMANTIC_URL,
            json={"name": name, "path": path},
            timeout=10.0,
        )
        resp.raise_for_status()
        data = resp.json()
        suggested = str(data.get("suggested_name", "")).strip()

        result = None
        if suggested and suggested != name:
            result = (name, suggested)

        # Сохраняем в кэш
        _LLM_CACHE[cache_key] = result
        return result

    except Exception as exc:
        # Если сервис недоступен — не ломаем основной пайплайн.
        # Сохраняем в кэш как None чтобы не пытаться снова
        _LLM_CACHE[cache_key] = None
        print(f"[semantic] не удалось получить рекомендацию: {exc}")
    return None


def process_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    cfg = CONFIG

    cat_cols: List[str] = [c for c in cfg["category_columns"] if c in df.columns]
    param_col = cfg["param_name_column"]
    value_col = cfg["param_value_column"]
    pid_col = cfg["param_id_column"]

    # Подсчитываем уникальные категории для LLM анализа
    if SEMANTIC_URL:
        unique_categories = set()
        for col in cat_cols:
            if col in df.columns:
                unique_vals = df[col].dropna().unique()
                unique_categories.update([str(v).strip().lower() for v in unique_vals if v])

        # Вычитаем уже закэшированные
        uncached = [c for c in unique_categories if c not in _LLM_CACHE]
        print(f"[LLM] Уникальных категорий: {len(unique_categories)}, уже в кэше: {len(_LLM_CACHE)}, нужно проверить: {len(uncached)}")

    text_cols: List[str] = []
    text_cols.extend(cat_cols)
    if param_col in df.columns:
        text_cols.append(param_col)
    if value_col in df.columns:
        text_cols.append(value_col)

    # создаём колонки для исправлений/комментариев
    for col in text_cols:
        df[f"{col}__correct"] = ""
        df[f"{col}__comment"] = ""

    # профиль части речи по параметрам
    print("[DEBUG] Строим профиль части речи...")
    param_pos_profile = build_param_pos_profile(df)
    print(f"[DEBUG] Профиль части речи построен: {len(param_pos_profile)} параметров")

    # профиль регистра по параметрам
    print("[DEBUG] Строим профиль регистра...")
    param_case_profile = build_param_case_profile(df)
    print(f"[DEBUG] Профиль регистра построен: {len(param_case_profile)} параметров")

    # профиль числа (singular/plural) по параметрам
    print("[DEBUG] Строим профиль числа...")
    param_number_profile = build_param_number_profile(df)
    print(f"[DEBUG] Профиль числа построен: {len(param_number_profile)} параметров")

    print(f"[DEBUG] Начинаем обработку {len(df)} строк...")
    for idx, row in df.iterrows():
        # категории — множ. число + орфография
        for col in cat_cols:
            val = row.get(col)
            if pd.isna(val) or val is None:
                continue
            text = str(val).strip()
            if not text:
                continue

            corrections: List[str] = []
            comments: List[str] = []

            format_issue = ensure_category_format(text)
            if format_issue:
                corrections.append(format_issue[1])
                comments.append("Формат категории: с заглавной буквы, именительный падеж, множественное число")

            spell_issue = check_spelling(text)
            if spell_issue:
                if not corrections:
                    corrections.append(spell_issue[1])
                comments.append("Орфография/грамматика")

            # Семантическая рекомендация от LLM-сервиса (с кэшированием)
            # Каждая уникальная категория проверяется только 1 раз
            if SEMANTIC_URL and not corrections:  # Используем LLM только если нет локальных ошибок
                full_path = " > ".join(
                    str(row.get(c)).strip()
                    for c in cat_cols
                    if c in df.columns and not pd.isna(row.get(c))
                )
                sem_issue = semantic_category_suggestion(text, full_path)
                if sem_issue:
                    corrections = [sem_issue[1]]
                    comments.append("Рекомендация LLM")

            if corrections:
                df.at[idx, f"{col}__correct"] = corrections[0]
                df.at[idx, f"{col}__comment"] = "; ".join(comments)

        # параметр и его значение
        param_name = ""
        if param_col in df.columns:
            val = row.get(param_col)
            if val is not None and not pd.isna(val):
                param_name = str(val).strip()

        param_value = ""
        if value_col in df.columns:
            val = row.get(value_col)
            if val is not None and not pd.isna(val):
                param_value = str(val).strip()

        pid = str(row.get(pid_col)) if pid_col in df.columns and not pd.isna(row.get(pid_col)) else None

        # орфография имени параметра
        if param_col in df.columns and param_name:
            spell_issue_param = check_spelling(param_name)
            if spell_issue_param:
                df.at[idx, f"{param_col}__correct"] = spell_issue_param[1]
                df.at[idx, f"{param_col}__comment"] = "Орфография/грамматика названия параметра"

        # значение параметра
        if value_col in df.columns and param_value:
            corrections: List[str] = []
            comments: List[str] = []

            # паттерн «Другой + название параметра» с морфологией
            other_issue = normalize_other_pattern(param_name, param_value)
            if other_issue:
                corrections.append(other_issue[1])
                comments.append('Шаблон "Другой/Другое/Другая/Другие + название параметра"')

            # проверка единообразия регистра ПЕРВОЙ БУКВЫ в рамках param_id
            if pid and pid in param_case_profile:
                expected_case = param_case_profile[pid]
                case_issue = check_case_consistency(param_value, expected_case)
                if case_issue:
                    if not corrections:
                        corrections.append(case_issue[1])
                    case_label = "строчная буква" if expected_case == "lowercase" else "заглавная буква"
                    comments.append(
                        f"Регистр первой буквы не соответствует большинству значений этого параметра (ожидается {case_label})"
                    )

            # орфография/грамматика значения
            spell_issue_val = check_spelling(param_value)
            if spell_issue_val:
                if not corrections:
                    corrections.append(spell_issue_val[1])
                comments.append("Орфография/грамматика значения параметра")

            # однородность части речи внутри параметра
            if pid and pid in param_pos_profile:
                target_pos = param_pos_profile[pid]
                actual_pos = detect_pos(param_value)
                if actual_pos and actual_pos != target_pos:
                    comments.append(
                        f"Часть речи значения отличается от большинства по этому параметру (ожидается {target_pos})"
                    )

            # однородность числа (singular/plural) внутри параметра
            if pid and pid in param_number_profile and not has_preposition(param_value):
                target_number = param_number_profile[pid]
                parsed = _first_parse(param_value)
                if parsed:
                    if "plur" in parsed.tag and "nomn" in parsed.tag:
                        actual_number = "plural"
                    elif "sing" in parsed.tag or "nomn" in parsed.tag:
                        actual_number = "singular"
                    else:
                        actual_number = None

                    if actual_number and actual_number != target_number:
                        number_label = "единственном числе" if target_number == "singular" else "множественном числе"
                        comments.append(
                            f"Число значения отличается от большинства значений этого параметра (ожидается в {number_label})"
                        )

            if corrections:
                df.at[idx, f"{value_col}__correct"] = corrections[0]
                df.at[idx, f"{value_col}__comment"] = "; ".join(comments)

    # Filter to only rows with corrections
    def has_corrections(row: pd.Series) -> bool:
        """Check if row has any corrections in any __correct column"""
        for col in df.columns:
            if col.endswith("__correct"):
                val = str(row[col]).strip()
                if val and val != "" and not pd.isna(row[col]):
                    return True
        return False

    # Apply filter to keep only rows with errors
    mask = df.apply(has_corrections, axis=1)
    df_filtered = df[mask].copy()

    logger.info(f"Filtered results: {len(df_filtered)} rows with errors out of {len(df)} total rows")

    return df_filtered


def write_with_highlight(df_processed: pd.DataFrame, output_path: str) -> None:
    """
    Пишем в Excel и подсвечиваем оригинальные ячейки, у которых есть исправление.
    Ошибки - светло-красным, исправления - светло-зелёным.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_processed.to_excel(writer, sheet_name="Checked", index=False)
        wb = writer.book
        ws = wb["Checked"]

        # Светло-красный для ошибок, светло-зелёный для исправлений
        error_fill = PatternFill(start_color="FFFFCDD2", end_color="FFFFCDD2", fill_type="solid")
        correction_fill = PatternFill(start_color="FFC8E6C9", end_color="FFC8E6C9", fill_type="solid")

        header = [cell.value for cell in ws[1]]
        col_index = {name: idx + 1 for idx, name in enumerate(header)}

        for row_idx in range(2, ws.max_row + 1):
            for base_col in header:
                if base_col is None:
                    continue
                name = str(base_col)
                if name.endswith("__correct") or name.endswith("__comment"):
                    continue

                correct_col = f"{name}__correct"
                if correct_col not in col_index:
                    continue

                correct_cell = ws.cell(row=row_idx, column=col_index[correct_col])
                if correct_cell.value not in (None, "", " "):
                    # Подсвечиваем оригинальную ячейку с ошибкой красным
                    base_cell = ws.cell(row=row_idx, column=col_index[name])
                    base_cell.fill = error_fill

                    # Подсвечиваем ячейку с исправлением зелёным
                    correct_cell.fill = correction_fill


def main() -> None:
    cfg = CONFIG
    input_path = cfg["input_file"]
    output_path = cfg["output_file"]

    if not os.path.exists(input_path):
        print(f"Файл не найден: {input_path}")
        return

    df = pd.read_csv(input_path, sep=cfg["sep"], encoding=cfg["encoding"], dtype=str)

    print("Колонки входного CSV:")
    print(list(df.columns))

    df_processed = process_dataframe(df)
    write_with_highlight(df_processed, output_path)

    print(f"Готово. Результат сохранён в: {output_path}")
    print("В Excel-листе `Checked` ищи подсвеченные ячейки и колонки `__correct` / `__comment`.")


if __name__ == "__main__":
    main()


